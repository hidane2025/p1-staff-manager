"""P1 Staff Manager E2E — CSVをExcel (.xlsx) に変換"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TEST_DIR = Path(__file__).resolve().parent


def csv_to_xlsx(csv_path: Path, xlsx_path: Path, sheet_name: str) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    row_count = 0
    with csv_path.open("r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            ws.append(row)
            row_count += 1
            if i == 0:
                # ヘッダー装飾
                for j, _ in enumerate(row, 1):
                    cell = ws.cell(row=1, column=j)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="333333")
                    cell.alignment = Alignment(horizontal="center")

    # 列幅自動調整（簡易）
    for j in range(1, ws.max_column + 1):
        letter = get_column_letter(j)
        max_len = 0
        for cell in ws[letter]:
            val = str(cell.value or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[letter].width = min(max_len * 1.8 + 2, 40)
    ws.freeze_panes = "A2"

    wb.save(xlsx_path)
    return row_count


def main() -> None:
    outputs = [
        (TEST_DIR / "01_staff_master.csv", TEST_DIR / "01_staff_master.xlsx", "スタッフマスタ"),
        (TEST_DIR / "02_shift_kyoto.csv", TEST_DIR / "02_shift_kyoto.xlsx", "京都2026シフト表"),
    ]
    for csv_path, xlsx_path, sheet in outputs:
        if not csv_path.exists():
            print(f"  ⚠️ 入力なし: {csv_path}")
            continue
        n = csv_to_xlsx(csv_path, xlsx_path, sheet)
        print(f"  ✅ {xlsx_path.name} ({n}行)")


if __name__ == "__main__":
    main()
